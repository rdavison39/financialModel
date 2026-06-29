"""
excel_reader.py

Generic Excel workbook reader.

Provides a lightweight wrapper around openpyxl that can be reused by
all brokerage importers.

Author:
    Ron Davison / ChatGPT
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from src.importers.worksheet_helper import WorksheetHelper


class ExcelReader:
    """
    Generic Excel workbook reader.
    """

    def __init__(
        self,
        filename: str | Path,
    ) -> None:
        """
        Initialize the reader.

        Parameters
        ----------
        filename
            Path to the Excel workbook.
        """

        self._filename = Path(filename)
        self._workbook: Workbook | None = None

    @property
    def filename(self) -> Path:
        """
        Return the workbook filename.
        """

        return self._filename

    def open(self) -> None:
        """
        Open the workbook.

        Raises
        ------
        FileNotFoundError
            If the workbook does not exist.
        """

        if not self._filename.exists():
            raise FileNotFoundError(
                f"Workbook not found: {self._filename}"
            )

        self._workbook = load_workbook(
            filename=self._filename,
            data_only=True,
        )

    @property
    def workbook(self) -> Workbook:
        """
        Return the loaded workbook.

        Raises
        ------
        RuntimeError
            If open() has not been called.
        """

        if self._workbook is None:
            raise RuntimeError(
                "Workbook has not been opened."
            )

        return self._workbook

    @property
    def sheet_names(self) -> list[str]:
        """
        Return the workbook's worksheet names.
        """

        return self.workbook.sheetnames

    @property
    def worksheet_count(self) -> int:
        """
        Return the number of worksheets.
        """

        return len(self.sheet_names)

    def get_worksheet(
        self,
        name: str,
    ) -> WorksheetHelper:
        """
        Return a worksheet by name.

        Raises
        ------
        KeyError
            If the worksheet does not exist.
        """

        if name not in self.sheet_names:
            raise KeyError(
                f"Worksheet '{name}' not found."
            )

        return WorksheetHelper(
            self.workbook[name]
        )

    @property
    def first_worksheet(self) -> WorksheetHelper:
        """
        Return the first worksheet.
        """

        return WorksheetHelper(
            self.workbook[self.sheet_names[0]]
        )

    def close(self) -> None:
        """
        Close the workbook.
        """

        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None

    def __enter__(self) -> "ExcelReader":
        """
        Context manager entry.
        """

        self.open()
        return self

    def __exit__(
        self,
        exc_type,
        exc_value,
        traceback,
    ) -> None:
        """
        Context manager exit.
        """

        self.close()