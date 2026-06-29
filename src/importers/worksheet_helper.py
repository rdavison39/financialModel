"""
worksheet_helper.py

Convenience wrapper around an openpyxl Worksheet.

Provides helper methods for reading worksheet values and searching
for text. This class contains no brokerage-specific logic and can
be reused by all importers.

Author:
    Ron Davison / ChatGPT
"""

from __future__ import annotations

from typing import Any

from openpyxl.worksheet.worksheet import Worksheet


class WorksheetHelper:
    """
    Convenience wrapper around an openpyxl worksheet.
    """

    def __init__(
        self,
        worksheet: Worksheet,
    ) -> None:
        """
        Initialize the helper.

        Parameters
        ----------
        worksheet
            OpenPyXL worksheet.
        """

        self._worksheet = worksheet

    @property
    def worksheet(self) -> Worksheet:
        """
        Return the wrapped worksheet.
        """

        return self._worksheet

    @property
    def max_row(self) -> int:
        """
        Return the maximum worksheet row.
        """

        return self._worksheet.max_row

    @property
    def max_column(self) -> int:
        """
        Return the maximum worksheet column.
        """

        return self._worksheet.max_column

    #
    # ------------------------------------------------------------------
    # Cell Access
    # ------------------------------------------------------------------
    #

    def cell_value(
        self,
        row: int,
        column: int,
    ) -> Any:
        """
        Return the raw cell value.

        Returns None if the cell is outside the worksheet.
        """

        if row < 1 or column < 1:
            return None

        if row > self.max_row or column > self.max_column:
            return None

        return self._worksheet.cell(
            row=row,
            column=column,
        ).value

    def cell_text(
        self,
        row: int,
        column: int,
    ) -> str:
        """
        Return the cell as stripped text.
        """

        value = self.cell_value(
            row,
            column,
        )

        if value is None:
            return ""

        return str(value).strip()

    #
    # ------------------------------------------------------------------
    # Row Helpers
    # ------------------------------------------------------------------
    #

    def row_values(
        self,
        row: int,
    ) -> list[Any]:
        """
        Return every value in a worksheet row.
        """

        return [
            self.cell_value(
                row,
                column,
            )
            for column in range(
                1,
                self.max_column + 1,
            )
        ]

    def row_text(
        self,
        row: int,
    ) -> list[str]:
        """
        Return every cell in a row as stripped text.
        """

        return [
            self.cell_text(
                row,
                column,
            )
            for column in range(
                1,
                self.max_column + 1,
            )
        ]

    def is_empty_row(
        self,
        row: int,
    ) -> bool:
        """
        Determine whether every cell in a row is empty.
        """

        return all(
            text == ""
            for text in self.row_text(row)
        )

    #
    # ------------------------------------------------------------------
    # Search Helpers
    # ------------------------------------------------------------------
    #

    @staticmethod
    def _normalize(
        text: str,
        case_sensitive: bool,
    ) -> str:
        """
        Normalize text for comparisons.
        """

        text = text.strip()

        if not case_sensitive:
            text = text.lower()

        return text

    def find_text(
        self,
        text: str,
        *,
        case_sensitive: bool = False,
    ) -> tuple[int, int] | None:
        """
        Find text anywhere in the worksheet.

        Returns
        -------
        tuple(row, column)
            If found.

        None
            If not found.
        """

        target = self._normalize(
            text,
            case_sensitive,
        )

        if target == "":
            return None

        for row in range(
            1,
            self.max_row + 1,
        ):
            for column in range(
                1,
                self.max_column + 1,
            ):
                candidate = self._normalize(
                    self.cell_text(
                        row,
                        column,
                    ),
                    case_sensitive,
                )

                if candidate and target in candidate:
                    return (
                        row,
                        column,
                    )

        return None

    def find_text_in_row(
        self,
        text: str,
        row: int,
        *,
        case_sensitive: bool = False,
    ) -> int | None:
        """
        Find text within a row.

        Returns the column number.
        """

        target = self._normalize(
            text,
            case_sensitive,
        )

        if target == "":
            return None

        for column in range(
            1,
            self.max_column + 1,
        ):
            candidate = self._normalize(
                self.cell_text(
                    row,
                    column,
                ),
                case_sensitive,
            )

            if candidate and target in candidate:
                return column

        return None

    def find_text_in_column(
        self,
        text: str,
        column: int,
        *,
        case_sensitive: bool = False,
    ) -> int | None:
        """
        Find text within a column.

        Returns the row number.
        """

        target = self._normalize(
            text,
            case_sensitive,
        )

        if target == "":
            return None

        for row in range(
            1,
            self.max_row + 1,
        ):
            candidate = self._normalize(
                self.cell_text(
                    row,
                    column,
                ),
                case_sensitive,
            )

            if candidate and target in candidate:
                return row

        return None